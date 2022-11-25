from csv import reader as csv_reader
from re import sub
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from string import ascii_uppercase
from typing import List


def custom_quit(msg: str) -> None:
    print(msg)
    quit()


def index_to_excel_letter(index: int):
    return ascii_uppercase[index]


class Translator:
    AZN: str = "Манаты"
    BYR: str = "Белорусские рубли"
    EUR: str = "Евро"
    GEL: str = "Грузинский лари"
    KGS: str = "Киргизский сом"
    KZT: str = "Тенге"
    RUR: str = "Рубли"
    UAH: str = "Гривны"
    USD: str = "Доллары"
    UZS: str = "Узбекский сум"
    currency_to_rub: {str, float} = {
        "Манаты": 35.68,
        "Белорусские рубли": 23.91,
        "Евро": 59.90,
        "Грузинский лари": 21.74,
        "Киргизский сом": 0.76,
        "Тенге": 0.13,
        "Рубли": 1,
        "Гривны": 1.64,
        "Доллары": 60.66,
        "Узбекский сум": 0.0055,
    }
    dataset_to_excel: {str, str} = {
        'Динамика уровня зарплат по годам: ': "Средняя зарплата",
        'Динамика количества вакансий по годам: ': "Количество вакансий",
        'Динамика уровня зарплат по годам для выбранной профессии: ': "Средняя зарплата - ",
        'Динамика количества вакансий по годам для выбранной профессии: ': "Количество вакансий - ",
        "Уровень зарплат по городам (в порядке убывания): ": "Уровень зарплат",
        "Доля вакансий по городам (в порядке убывания): ": "Доля вакансий"

    }

    def translate(self, key: str, dict_name: str = None) -> str:
        if dict_name is not None:
            return self.__getattribute__(dict_name)[key]
        return self.__getattribute__(key)

    def translate_currency_to_rub(self, currency: str) -> int or float:
        return self.currency_to_rub[currency]


class UserInterface:
    file_name: str
    profession_name: str

    def __init__(self):
        self.file_name = '../vacancies_by_year.csv'  # input('Введите название файла: ')
        self.profession_name = 'Программист'  # input('Введите название профессии: ')


class CSV:
    data: csv_reader
    title: list
    rows: list

    def __init__(self, file_name: str):
        with open(file_name, 'r', newline='', encoding='utf-8-sig') as file:
            self.data = csv_reader(file)
            try:
                self.title = next(self.data)
            except StopIteration:
                custom_quit('Пустой файл')

            self.rows = [row for row in self.data
                         if len(list(filter(lambda word: word != '', row))) == len(self.title)]

            if len(self.rows) == 0:
                custom_quit('Нет данных')


class Salary:
    salary_from: int
    salary_to: int
    salary_currency: str

    def set_field(self, key: str, value: str):
        if key == 'salary_currency':
            value = translator.translate(value)
        if key in ['salary_from', 'salary_to']:
            value = float(value)
        self.__setattr__(key, value)

    def get_average_in_rur(self) -> int:
        return int(translator.translate_currency_to_rub(self.salary_currency) *
                   (float(self.salary_from) + float(self.salary_to)) // 2)


class Vacancy:
    name: str
    salary: Salary
    areaName: str
    published_at: int

    def __init__(self, fields: dict):
        for key, value in fields.items():
            if not self.check_salary(key, value):
                self.__setattr__(key, self.get_correct_field(key, value))

    def get_field(self, field: str):
        if field in 'salary':
            return self.salary.get_average_in_rur()
        return self.__getattribute__(field)

    def check_salary(self, key: str, value: str) -> bool:
        is_salary = False
        if key in ['salary_from', 'salary_to', 'salary_currency']:
            if not hasattr(self, 'salary'):
                self.salary = Salary()
            self.salary.set_field(key, value)
            is_salary = True
        return is_salary

    @staticmethod
    def get_correct_field(key: str, value: str or list) -> int or str:
        if key == 'published_at':
            big, small = value[:19].split('T')
            year, month, day = big.split('-')
            # hours, minutes, seconds = small.split(':')
            return int(year)
        else:
            return value


class DataSet:
    profession_name: str
    profession_count: int
    vacancies: List[Vacancy]
    salary_by_years: {int, list}
    vacancies_by_years: {int, int}
    profession_salary_by_years: {int, list}
    profession_vacancies_by_years: {int, int}
    salaries_by_cities: {str, list}
    ratio_vacancy_by_cities: {str, float}
    city_vacancies_count: {str, int}

    def __init__(self, vacs: list, prof_name: str):
        self.profession_name = prof_name
        self.profession_count = 0
        self.vacancies = vacs
        self.salary_by_years = {}
        self.vacancies_by_years = {}
        self.profession_salary_by_years = {}
        self.profession_vacancies_by_years = {}
        self.salaries_by_cities = {}
        self.ratio_vacancy_by_cities = {}
        self.city_vacancies_count = {}

        self.get_data()

    def get_data(self) -> None:
        for vac in self.vacancies:
            self.process_vacancies_count('city_vacancies_count', 'area_name', vac)
        for vac in self.vacancies:
            self.process_salary('salary_by_years', 'published_at', vac)
            self.process_vacancies_count('vacancies_by_years', 'published_at', vac)
            if self.profession_name in vac.name:
                self.profession_count += 1
                self.process_salary('profession_salary_by_years', 'published_at', vac)
                self.process_vacancies_count('profession_vacancies_by_years', 'published_at', vac)
            self.process_salary('salaries_by_cities', 'area_name', vac)
            self.process_vacancies_count('ratio_vacancy_by_cities', 'area_name', vac)

        self.set_correct_cities_data()

    def process_salary(self, dict_name: str, field: str, vac: Vacancy) -> None:
        d = self.__getattribute__(dict_name)
        f = vac.get_field(field)
        if f not in d.keys():
            d[f] = [vac.salary.get_average_in_rur(), 1]
        else:
            d[f][0] += vac.salary.get_average_in_rur()
            d[f][1] += 1

    def process_vacancies_count(self, dict_name: str, field: str, vac: Vacancy) -> None:
        d = self.__getattribute__(dict_name)
        f = vac.get_field(field)
        if f not in d.keys():
            d[f] = 1
        else:
            d[f] += 1

    def set_correct_cities_data(self):
        for key, value in self.ratio_vacancy_by_cities.items():
            self.ratio_vacancy_by_cities[key] = round(value / len(self.vacancies), 4)

        d1 = dict(sorted(self.salaries_by_cities.items(), key=lambda i: i[1][1] / i[1][0]))
        self.salaries_by_cities = self.get_first_ten_correct(d1)

        d2 = dict(sorted(self.ratio_vacancy_by_cities.items(), key=lambda i: i[1], reverse=True))
        self.ratio_vacancy_by_cities = self.get_first_ten_correct(d2)

    def get_first_ten_correct(self, d: dict) -> dict:
        count = 0
        res = {}
        for key, value in d.items():
            if count == 10:
                break
            if self.city_vacancies_count[key] >= len(self.vacancies) // 100:
                res[key] = value
                count += 1
        return res

    def get_statistics(self) -> list:
        res = []
        to_print: {str, dict} \
            = {"Динамика уровня зарплат по годам: ": self.salary_by_years,
               "Динамика количества вакансий по годам: ": self.vacancies_by_years,
               "Динамика уровня зарплат по годам для выбранной профессии: ": self.profession_salary_by_years,
               "Динамика количества вакансий по годам для выбранной профессии: ": self.profession_vacancies_by_years,
               "Уровень зарплат по городам (в порядке убывания): ": self.salaries_by_cities,
               "Доля вакансий по городам (в порядке убывания): ": self.ratio_vacancy_by_cities}
        for key, value in to_print.items():
            if len(value) == 0:
                value = {k: 0 for k in self.salary_by_years.keys()}
            for k, v in value.items():
                if type(v) is list:
                    value[k] = v[0] // v[1]
            res.append({key: value})
        return res


def parse_html(line: str) -> str:
    line = sub('<.*?>', '', line)
    res = [' '.join(word.split()) for word in line.replace("\r\n", "\n").split('\n')]
    return res[0] if len(res) == 1 else res  # Спасибо Яндекс.Контесту за еще один костыль!


def parse_row_vacancy(row_vacs: list) -> dict:
    return dict(zip(title, map(parse_html, row_vacs)))


class Report:
    workbook: Workbook
    data: list

    def __init__(self, data: list):
        self.workbook = Workbook()
        self.data = data

    def generate_excel(self):
        self.fill_with_statistics()
        self.workbook.save('report2.xlsx')

    def fill_with_statistics(self):
        # self.fill_salaries_statistics()
        self.fill_cities_statistics()

    def fill_salaries_statistics(self):
        ws = self.workbook.active
        ws.title = 'Статистика по годам'
        years = [k for k in list(self.data[0].items())[0][1].keys()]
        self.fill_column(ws, 'Год', years, [col for col in ws.iter_cols(min_row=1, max_row=17, max_col=1)][0])
        unordered_data = [self.data[0], self.data[2], self.data[1], self.data[3]]
        for column, cust_col in zip(ws.iter_cols(min_row=1, min_col=2, max_col=5, max_row=17), unordered_data):
            t = list(cust_col.items())[0]
            (key, dictionary) = t[0], t[1]
            header = translator.translate(key, "dataset_to_excel")
            actual_values = [v for v in dictionary.values()]
            self.fill_column(ws, header, actual_values, column)

        self.update_worksheet_settings(ws)


    def fill_cities_statistics(self):
        self.workbook.create_sheet("Статистика по городам")
        ws = self.workbook["Статистика по городам"]
        a = list(self.data[4].items())[0]
        (first_header, first_dictionary) = translator.translate(a[0], "dataset_to_excel"), a[1]
        first_cities = [k for k in first_dictionary.keys()]
        first_values = [v for v in first_dictionary.values()]

        self.fill_column(ws, 'Город', first_cities, [cell[0] for cell in ws['A1':'A11']])
        self.fill_column(ws, first_header, first_values, [cell[0] for cell in ws['B1': 'B11']])

        b = list(self.data[5].items())[0]
        (second_header, second_dictionary) = translator.translate(b[0], "dataset_to_excel"), b[1]
        second_cities = [k for k in second_dictionary.keys()]
        second_values = [v for v in second_dictionary.values()]

        self.fill_column(ws, 'Город', second_cities, [cell[0] for cell in ws['D1':'D11']])
        self.fill_column(ws, second_header, second_values, [cell[0] for cell in ws['E1': 'E11']])

        self.set_column_percent(ws, [cell[0] for cell in ws['E1': 'E11']])
        self.update_worksheet_settings(ws)

    def fill_column(self, ws, header: str, data: list, column_cells: list):
        if '-' in header:
            header += ds.profession_name
        column_cells[0].value = header
        for cell, value in zip(column_cells[1:], data):
            cell.value = value

    def set_column_percent(self, ws, column: list):
        for cell in column:
            cell.number_format = FORMAT_PERCENTAGE_00

    def update_worksheet_settings(self, ws):
        self.set_borders(ws)
        self.set_column_width(ws)

    def set_borders(self, ws):
        isFirstRow = True
        for row in ws.rows:
            for cell in row:
                if not cell.value:
                    continue
                cell.border = Border(top=Side(border_style="thin", color="000000"),
                                     left=Side(border_style="thin", color="000000"),
                                     right=Side(border_style="thin", color="000000"),
                                     bottom=Side(border_style="thin", color="000000"))
                if isFirstRow:
                    cell.font = Font(bold=True)
            isFirstRow = False

    def set_column_width(self, ws):
        a = {0: "A", 1: "B", 2: "C", 3: "D", 4: "E", 6: "F", 7: "G"}
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value)) + 1))

        for col, value in dims.items():
            ws.column_dimensions[a[col - 1]].width = value


if __name__ == '__main__':
    translator = Translator()
    ui = UserInterface()
    csv = CSV(ui.file_name)
    title, row_vacancies = csv.title, csv.rows
    vacancies = [Vacancy(parse_row_vacancy(row_vac)) for row_vac in row_vacancies]
    ds = DataSet(vacancies, ui.profession_name)
    statistics = ds.get_statistics()
    report = Report(statistics)
    report.generate_excel()
